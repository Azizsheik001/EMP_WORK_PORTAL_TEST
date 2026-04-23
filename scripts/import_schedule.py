#!/usr/bin/env python3
"""
Import employee schedules from Excel into the AGS Workforce database via API.
Reads All_Team_Schedule_April5th.xlsx and creates shift assignments for 4 weeks.
"""

import re
import sys
import json
import urllib.request
import urllib.error
import openpyxl
from datetime import date, timedelta

BASE_URL = "http://localhost:3000"
EXCEL_PATH = "docs/All_Team_Schedule_April5th.xlsx"
SHEET_NAME = "All team schedule"

# Week 1 starts Mon Mar 9, 2026. Repeat for 4 weeks.
WEEK1_START = date(2026, 3, 9)
NUM_WEEKS = 4

# Day offsets: Mon=0, Tue=1, ..., Sun=6 (columns D-J)
DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def api_request(method, path, body=None, token=None):
    """Make an HTTP request to the backend API."""
    url = f"{BASE_URL}{path}"
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    data = json.dumps(body).encode() if body else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        error_body = e.read().decode()
        print(f"  API Error {e.code}: {error_body}")
        raise


def login(email, password):
    """Login and return auth token."""
    result = api_request("POST", "/api/auth/login", {"email": email, "password": password})
    return result["token"]


def normalize_time(t):
    """Normalize a time string like '1:30' -> '01:30', '04:30' -> '04:30'."""
    t = t.strip()
    if ":" in t:
        parts = t.split(":")
        h = parts[0].zfill(2)
        m = parts[1].zfill(2)
        return f"{h}:{m}"
    return t


def parse_shift(raw):
    """
    Parse a shift string like '09:00 - 18:00', '16:30 - 1:30', '08:30 -17:30',
    '14:30 \u2013 23:30' (en-dash), 'OFF'.
    Returns (start_time, end_time, is_off).
    """
    if not raw:
        return None, None, True
    raw = str(raw).strip()
    if raw.upper() == "OFF":
        return None, None, True
    # Split on hyphen or en-dash, with optional surrounding spaces
    # Use regex to split on ' - ', ' -', '- ', '-', ' \u2013 ', '\u2013'
    parts = re.split(r'\s*[\-\u2013]\s*', raw, maxsplit=1)
    if len(parts) != 2:
        print(f"  WARNING: Could not parse shift '{raw}', treating as OFF")
        return None, None, True
    start = normalize_time(parts[0])
    end = normalize_time(parts[1])
    return start, end, False


def read_excel(path):
    """Read the Excel file and return list of (employee_name, employee_id, shifts[7])."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME]
    employees = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name = row[1]  # Column B
        emp_id = row[2]  # Column C
        if not name:
            continue
        name = str(name).strip()
        shifts = []
        for i in range(7):  # Columns D-J (indices 3-9)
            cell_val = row[3 + i] if (3 + i) < len(row) else None
            shifts.append(cell_val)
        employees.append((name, emp_id, shifts))
    return employees


def match_employees(excel_employees, api_users):
    """
    Match Excel employee names to API user IDs.
    Returns dict: excel_name -> user_id, and list of unmatched names.
    """
    # Build lookup by normalized name
    user_lookup = {}
    for u in api_users:
        normalized = u["name"].strip().lower()
        user_lookup[normalized] = u["id"]

    matched = {}
    unmatched = []
    for name, emp_id, shifts in excel_employees:
        norm = name.strip().lower()
        if norm in user_lookup:
            matched[name] = user_lookup[norm]
        else:
            # Try partial matching (last name, first name variations)
            found = False
            for api_name, uid in user_lookup.items():
                # Check if all words in one name appear in the other
                excel_words = set(norm.split())
                api_words = set(api_name.split())
                if excel_words == api_words:
                    matched[name] = uid
                    found = True
                    break
            if not found:
                unmatched.append(name)
    return matched, unmatched


def build_assignments(excel_employees, name_to_uid, week_offset):
    """
    Build assignment list for a given week (0-based offset from WEEK1_START).
    Returns list of assignment dicts for the bulk API.
    """
    week_start = WEEK1_START + timedelta(weeks=week_offset)
    assignments = []
    for name, emp_id, shifts in excel_employees:
        uid = name_to_uid.get(name)
        if not uid:
            continue
        for day_idx in range(7):
            shift_date = week_start + timedelta(days=day_idx)
            start, end, is_off = parse_shift(shifts[day_idx])
            assignment = {
                "user_id": uid,
                "shift_date": shift_date.isoformat(),
                "is_off": is_off,
            }
            if not is_off:
                assignment["shift_start_time"] = start
                assignment["shift_end_time"] = end
            assignments.append(assignment)
    return assignments


def main():
    print("=" * 60)
    print("AGS Workforce Schedule Import")
    print("=" * 60)

    # Step 1: Read Excel
    print("\n[1] Reading Excel file...")
    excel_employees = read_excel(EXCEL_PATH)
    print(f"    Found {len(excel_employees)} employees in Excel")

    # Step 2: Login
    print("\n[2] Logging in as admin...")
    token = login("admin@amgsol.com", "admin123")
    print("    Login successful")

    # Step 3: Get users
    print("\n[3] Fetching users from API...")
    users_resp = api_request("GET", "/api/users", token=token)
    api_users = users_resp.get("users", users_resp) if isinstance(users_resp, dict) else users_resp
    # Handle if response is { users: [...] } or just [...]
    if isinstance(api_users, dict) and "users" in api_users:
        api_users = api_users["users"]
    print(f"    Found {len(api_users)} users in database")

    # Step 4: Get clients
    print("\n[4] Fetching clients from API...")
    clients_resp = api_request("GET", "/api/clients", token=token)
    clients = clients_resp.get("clients", clients_resp) if isinstance(clients_resp, dict) else clients_resp
    if isinstance(clients, dict) and "clients" in clients:
        clients = clients["clients"]
    print(f"    Found {len(clients)} clients:")
    for c in clients:
        print(f"      - {c['name']} (id: {c['id']})")
    # Use the first client
    client_id = clients[0]["id"]
    print(f"    Using client: {clients[0]['name']} ({client_id})")

    # Step 5: Match employees
    print("\n[5] Matching employees...")
    name_to_uid, unmatched = match_employees(excel_employees, api_users)
    print(f"    Matched: {len(name_to_uid)} / {len(excel_employees)}")
    if unmatched:
        print(f"    WARNING - {len(unmatched)} employees NOT found in database:")
        for name in unmatched:
            print(f"      - {name}")

    # Step 6: Upload shifts for each week
    print("\n[6] Uploading shifts...")
    total_created = 0
    for week in range(NUM_WEEKS):
        week_start = WEEK1_START + timedelta(weeks=week)
        week_end = week_start + timedelta(days=6)
        print(f"\n    Week {week + 1}: {week_start} to {week_end}")

        assignments = build_assignments(excel_employees, name_to_uid, week)
        print(f"      Assignments to upload: {len(assignments)}")

        if not assignments:
            print("      Skipping (no assignments)")
            continue

        # The bulk endpoint may have limits, send in batches of 200
        BATCH_SIZE = 200
        week_count = 0
        for i in range(0, len(assignments), BATCH_SIZE):
            batch = assignments[i:i + BATCH_SIZE]
            body = {
                "client_id": client_id,
                "assignments": batch,
            }
            result = api_request("POST", "/api/shifts/bulk", body=body, token=token)
            created = result.get("count", len(batch))
            week_count += created

        total_created += week_count
        print(f"      Created/updated: {week_count} assignments")

    # Summary
    print("\n" + "=" * 60)
    print("IMPORT COMPLETE")
    print(f"  Total assignments created/updated: {total_created}")
    print(f"  Employees matched: {len(name_to_uid)}")
    print(f"  Employees unmatched: {len(unmatched)}")
    if unmatched:
        print(f"  Unmatched names: {', '.join(unmatched)}")
    print("=" * 60)


if __name__ == "__main__":
    main()
